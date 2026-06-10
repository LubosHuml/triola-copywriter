import openpyxl
import re

file_path = "triola_marketing_data.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Searching for sheets with sales arguments or detailed text...")
print(f"Total sheets: {len(wb.sheetnames)}")

found_sheets = []

for name in wb.sheetnames:
    ws = wb[name]
    # Check first row headers
    headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
    
    # Check if 'argument' or 'prodejní' is in headers
    matched_cols = []
    for idx, h in enumerate(headers, 1):
        if 'argument' in h or 'prodejn' in h:
            matched_cols.append((idx, h))
            
    if matched_cols:
        print(f"\nSheet '{name}' has sales arguments columns:")
        for idx, col_name in matched_cols:
            print(f" - Col {idx} ({openpyxl.utils.get_column_letter(idx)}): '{col_name}'")
            # Let's show some non-empty samples from this column
            samples = []
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=idx).value
                if val and str(val).strip() and str(val).lower() != 'x':
                    samples.append((r, str(val)[:60]))
                if len(samples) >= 3:
                    break
            if samples:
                print("   Samples:")
                for r, s in samples:
                    print(f"     Row {r}: {s}")
            else:
                print("   (No detailed samples, mostly empty or 'x')")
                
        found_sheets.append(name)

# Let's inspect some rows from sheets with 'Triola' in name to check their columns
print("\nScanning Triola sheets for model codes and columns:")
triola_sheets = [s for s in wb.sheetnames if 'triola' in s.lower()]
for ts in triola_sheets:
    ws = wb[ts]
    headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
    # Check if there is a column for model or střih
    has_model = any('kód' in h.lower() or 'číslo' in h.lower() or 'nomenklatura' in h.lower() or 'fazón' in h.lower() for h in headers)
    print(f" - Sheet '{ts}': Columns={len(headers)} | Headers={headers[:8]}... | Has Model Col={has_model}")
    
    # Print a few samples of rows to understand where the arguments are
    if ts != "Triola stálá kolekce":
        print("   Sample row 2:")
        row2 = [str(cell.value)[:40] if cell.value is not None else "" for cell in ws[2]]
        print(f"     Row 2: {row2[:8]}...")
