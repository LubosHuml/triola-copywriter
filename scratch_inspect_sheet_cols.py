import openpyxl

file_path = "triola_marketing_data.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

inspect_sheets = ["Triola stálá kolekce", "Triola JL 26", "Triola Plavky 2026"]

for sheet_name in inspect_sheets:
    if sheet_name in wb.sheetnames:
        print(f"\n==================================================")
        print(f"SHEET: {sheet_name}")
        print(f"==================================================")
        ws = wb[sheet_name]
        
        # Print first 6 rows, first 10 columns
        for row_idx, row in enumerate(ws.iter_rows(max_row=8, max_col=12, values_only=True), 1):
            # Print non-empty row values
            vals = [str(x)[:30] if x is not None else "" for x in row]
            # Print only if row is not entirely empty
            if any(vals):
                print(f"Row {row_idx}: {vals}")
    else:
        print(f"Sheet '{sheet_name}' not found.")
