import os
import openpyxl
import re
import logging
import csv

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

COLOR_CODE_TO_NAME = {
    "01": "bílá - puntíček",
    "02": "stříbrná",
    "03": "bílá",
    "04": "černá",
    "05": "modrá",
    "06": "nugát",
    "07": "čokoládová",
    "08": "zelená",
    "09": "šedá",
    "10": "vínová",
    "11": "hnědá",
    "12": "slonová kost",
    "13": "jeans",
    "14": "rosé",
    "15": "zlatozelená",
    "16": "smetanovo - modrá",
    "17": "fialová",
    "18": "khaki",
    "19": "černo - zelená",
    "20": "jahodová",
    "21": "červená",
    "22": "modro - růžová",
    "23": "průhledná",
    "24": "make-up",
    "25": "světle modrá",
    "26": "ash",
    "27": "tyrkysová",
    "28": "světle zelená",
    "29": "fuchsiová",
    "30": "perlová",
    "31": "shadow",
    "32": "cherry",
    "33": "malinová",
    "34": "modro - šedá",
    "35": "navy",
    "36": "modro-červená",
    "37": "medová",
    "38": "šedo - fialová",
    "39": "petrolejová",
    "40": "lilková",
    "41": "lunar rock",
    "42": "smetanová kvítek",
    "43": "cihlová",
    "44": "lososová",
    "45": "černo-červená",
    "46": "šedo - modrá",
    "47": "dusty rose",
    "48": "perleťově modrá",
    "49": "olivová",
    "50": "zlatá",
    "51": "karmínová",
    "52": "šedo - růžová",
    "53": "tmavě šedá",
    "54": "koňaková",
    "55": "měděná",
    "67": "deco rose",
    "68": "cappuccino",
    "69": "bílo - zelená",
    "70": "žlutá",
    "71": "mint",
    "72": "smaragdová",
    "77": "šedo - hnědá",
    "78": "ametystová",
    "79": "nachová",
    "80": "oranžová",
    "81": "růžová",
    "82": "tm. růžová",
    "83": "smetana",
    "84": "champagne",
    "85": "eurová",
    "86": "tělová",
    "87": "dračí ovoce",
    "88": "bordó",
    "89": "lambrusco",
    "90": "černo - bílá",
    "91": "limeta",
    "92": "pudr",
    "93": "světle šedá",
    "95": "béžová",
    "97": "černá - tisk",
    "98": "kardinál",
    "99": "vícebarevná"
}

def resolve_color(model_code, color_code, products_db):
    """
    Attempts to match a color code (like 88 or 04) to a standard Czech color name.
    Prioritizes the static code mapping, then matches against feed variants,
    falling back to feed colors if the code is unknown.
    """
    if not color_code:
        return "neuvedena"
        
    color_code_str = str(color_code).strip().zfill(2) # Normalise e.g. "4" -> "04", "88" -> "88"
    expected_color = COLOR_CODE_TO_NAME.get(color_code_str)
    
    # 1. Check if model is in products db
    if model_code in products_db:
        feed_colors = products_db[model_code].get("all_colors", [])
        if expected_color:
            # Check if there is an exact or substring match in feed colors
            for fc in feed_colors:
                if expected_color in fc.lower() or fc.lower() in expected_color:
                    return fc
            # If not in feed but we have a name, return the mapped name (new variant)
            return expected_color
        else:
            # If we don't have expected color, fall back to single feed color if it exists
            if len(feed_colors) == 1:
                return feed_colors[0]
                
    # 2. Fallback to static mapping or raw code
    return expected_color if expected_color else f"kód {color_code}"

def find_header_row_and_mapping(ws):
    """
    Prohledá prvních 10 řádků listu (openpyxl) a najde záhlaví.
    Deleguje na detect_header_mapping, aby Excel i Google Sheets pouzivaly stejna pravidla.
    """
    last = min(ws.max_row, 70)
    rows = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, last + 1)]
    return detect_header_mapping(rows)


def detect_header_mapping(rows):
    """
    rows: list of lists (0-indexed radky, hodnoty bunky nebo None).
    Vrací (header_row_num 1-based, col_map s 0-based indexy sloupcu).
    Jediny zdroj pravdy pro nazvy sloupcu - pouziva Excel import i Google Sheets.
    """
    n_cols = max((len(r) for r in rows), default=0)

    def cell(r_idx, c_idx):
        """1-based radek, 0-based sloupec."""
        if r_idx - 1 >= len(rows):
            return None
        row = rows[r_idx - 1]
        return row[c_idx] if c_idx < len(row) else None

    col_map = {
        "code": -1,
        "cut": -1,
        "color": -1,
        "collection": -1,
        "arguments": -1,
        "product_name": -1,
        "design_name": -1,
        "brand": -1,
        "material": -1,
        "size": -1,
        "eshop_name": -1,
        "short_name": -1,
        "eshop_desc1": -1,
        "eshop_desc2": -1,
        "meta_title": -1,
        "meta_desc": -1,
    }
    
    best_row = 1
    best_score = -1
    best_mapping = {}
    
    for r in range(1, min(11, len(rows) + 1)):
        row_vals = []
        for c in range(n_cols):
            val = cell(r, c)
            row_vals.append(str(val).strip() if val is not None else "")
            
        score = 0
        current_map = {k: -1 for k in col_map}
        
        for idx, val in enumerate(row_vals):
            val_lower = val.lower()
            if not val_lower:
                continue
                
            # Code mapping (nomenklatura ma prednost - byva to skutecny kod produktu)
            if "nomenklatura" in val_lower:
                current_map["code"] = idx
                score += 6
            elif "fazón" in val_lower or "číslo" in val_lower or "kod" in val_lower or val_lower == "code":
                if current_map["code"] == -1:
                    current_map["code"] = idx
                    score += 5
            # Typ produktu (sloupec "PRODUKT": podprsenka/kalhotky/osuška...)
            elif val_lower == "produkt" or "typ produktu" in val_lower:
                if current_map["product_name"] == -1:
                    current_map["product_name"] = idx
                    score += 2
            # Znacka (sassa, Triola...)
            elif "značka" in val_lower or "znacka" in val_lower or val_lower == "brand":
                if current_map["brand"] == -1:
                    current_map["brand"] = idx
                    score += 2
            # Nazev designu/kolekce vyrobce (ne e-shop nazev, ne SK mutace)
            elif ("název" in val_lower or "nazev" in val_lower) and "shop" not in val_lower and " sk" not in val_lower and "kolekce" not in val_lower and "krátký" not in val_lower and "kratky" not in val_lower and "meta" not in val_lower:
                if current_map["design_name"] == -1:
                    current_map["design_name"] = idx
                    score += 2
            elif "materiál" in val_lower or "material" in val_lower:
                if current_map["material"] == -1:
                    current_map["material"] = idx
                    score += 1
            elif "velikost" in val_lower:
                if current_map["size"] == -1:
                    current_map["size"] = idx
                    score += 1
            # Arguments mapping
            elif "argument" in val_lower or "prodejní" in val_lower or "prodejni" in val_lower:
                if current_map["arguments"] == -1:
                    current_map["arguments"] = idx
                    score += 4
            # Střih mapping
            elif "střih" in val_lower or "strih" in val_lower:
                if current_map["cut"] == -1:
                    current_map["cut"] = idx
                    score += 2
            # Color mapping
            elif "barva" in val_lower or val_lower == "color":
                if current_map["color"] == -1:
                    current_map["color"] = idx
                    score += 2
            # Collection mapping
            elif "kolekce" in val_lower or "název kolekce" in val_lower:
                if current_map["collection"] == -1:
                    current_map["collection"] = idx
                    score += 2
            # Outputs mapping
            elif ("krátký název" in val_lower or "kratky nazev" in val_lower) and not val_lower.endswith(" sk"):
                if current_map["short_name"] == -1:
                    current_map["short_name"] = idx
                    score += 3
            elif ("e-shop název" in val_lower or "eshop nazev" in val_lower) and not val_lower.endswith(" sk"):
                if current_map["eshop_name"] == -1:
                    current_map["eshop_name"] = idx
                    score += 3
            elif "triola eshop popis 2" in val_lower or "popis 2" in val_lower:
                if current_map["eshop_desc2"] == -1:
                    current_map["eshop_desc2"] = idx
                    score += 3
            elif "triola eshop popis" in val_lower or "popis 1" in val_lower or "eshop popis" in val_lower:
                if current_map["eshop_desc1"] == -1:
                    current_map["eshop_desc1"] = idx
                    score += 3
            elif "meta title" in val_lower or "eshop meta title" in val_lower:
                if current_map["meta_title"] == -1:
                    current_map["meta_title"] = idx
                    score += 3
            elif "meta description" in val_lower or "eshop meta description" in val_lower or "meta desc" in val_lower:
                if current_map["meta_desc"] == -1:
                    current_map["meta_desc"] = idx
                    score += 3
                    
        if score > best_score:
            best_score = score
            best_row = r
            best_mapping = current_map
            
    # Fallback if code column not identified
    if best_score <= 0 or best_mapping.get("code", -1) == -1:
        best_row = 1
        best_mapping = {
            "code": 0,
            "cut": -1,
            "color": -1,
            "collection": -1,
            "arguments": 1,
            "product_name": -1,
            "design_name": -1,
            "brand": -1,
            "material": -1,
            "size": -1,
            "eshop_name": -1,
            "short_name": -1,
            "eshop_desc1": -1,
            "eshop_desc2": -1,
            "meta_title": -1,
            "meta_desc": -1,
        }
        
    # VALIDACE DAT: header muze obsahovat vic sloupcu se slovem "kod" (napr. prazdny
    # "Kod výrobce" vedle plne "Nomenklatura"). Zvoleny sloupec musi mit data pod sebou;
    # kdyz nema, vyber ze vsech kandidatu ten s nejvice neprazdnymi hodnotami.
    def _data_count(col_idx):
        cnt = 0
        for rr in range(best_row + 1, min(best_row + 60, len(rows)) + 1):
            v = cell(rr, col_idx)
            if v is not None and str(v).strip() not in ("", "None"):
                cnt += 1
        return cnt

    if best_mapping.get("code", -1) != -1 and _data_count(best_mapping["code"]) == 0:
        code_keywords = ("nomenklatura", "fazón", "fazon", "číslo", "cislo", "kod", "code")
        candidates = []
        for idx2 in range(n_cols):
            hv = cell(best_row, idx2)
            hv_lower = str(hv).strip().lower() if hv is not None else ""
            if hv_lower and any(k in hv_lower for k in code_keywords):
                candidates.append((_data_count(idx2), idx2))
        candidates.sort(reverse=True)
        if candidates and candidates[0][0] > 0:
            logging.info(f"Sloupec kodu '{cell(best_row, best_mapping['code'])}' je prazdny - prepinam na sloupec {candidates[0][1]} ('{cell(best_row, candidates[0][1])}') s {candidates[0][0]} hodnotami.")
            best_mapping["code"] = candidates[0][1]

    return best_row, best_mapping

def parse_batch_excel(file_path, products_db):
    """
    Parses the uploaded Excel file. Skips empty rows.
    Automatically resolves the color names and checks if the model is in the DB.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Soubor {file_path} neexistuje.")
        
    logging.info(f"Načítání Excelu: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    header_row, mapping = find_header_row_and_mapping(ws)
    logging.info(f"Nalezen řádek záhlaví {header_row} s mapováním: {mapping}")
    
    rows = []
    # Loop from row header_row + 1
    for r in range(header_row + 1, ws.max_row + 1):
        raw_code = ws.cell(row=r, column=mapping["code"] + 1).value
        
        # Check if the code is None
        if raw_code is None:
            continue
            
        code_str = str(raw_code).strip()
        if not code_str or code_str == "None":
            continue
            
        if code_str.endswith(".0"):
            code_str = code_str[:-2]
            
        # Parse model and color code (e.g. 22859/88)
        parts = code_str.split('/')
        model_code = parts[0].strip()
        if model_code.endswith(".0"):
            model_code = model_code[:-2]
            
        color_code = parts[1].strip() if len(parts) > 1 else ""
        if color_code.endswith(".0"):
            color_code = color_code[:-2]
            
        # Check if there is a separate color column and color_code is not yet parsed
        color_name_override = None
        if not color_code and mapping["color"] != -1:
            raw_color = ws.cell(row=r, column=mapping["color"] + 1).value
            if raw_color is not None:
                color_str = str(raw_color).strip()
                if color_str.endswith(".0"):
                    color_str = color_str[:-2]
                if color_str.isdigit():
                    color_code = color_str
                else:
                    color_code = ""
                    color_name_override = color_str
        
        if color_name_override:
            color_name = color_name_override
        else:
            color_name = resolve_color(model_code, color_code, products_db)
            
        # Arguments
        args = ""
        if mapping["arguments"] != -1:
            raw_args = ws.cell(row=r, column=mapping["arguments"] + 1).value
            args = str(raw_args).strip() if raw_args is not None else ""
            
        # Volitelne sloupce: nazev produktu, material, velikost (pro nepradlove produkty)
        def _cell(key):
            if mapping.get(key, -1) == -1:
                return ""
            v = ws.cell(row=r, column=mapping[key] + 1).value
            return str(v).strip() if v is not None else ""

        product_name = _cell("product_name")
        design_name = _cell("design_name")
        brand = _cell("brand")
        material = _cell("material")
        size = _cell("size")

        # Heuristika znacky podle prefixu kodu, pokud sloupec Znacka chybi
        if not brand:
            code_upper = model_code.upper()
            for prefix, brand_name in (("SAS", "Sassa"),):
                if code_upper.startswith(prefix):
                    brand = brand_name
                    break

        # Pokud typ produktu chybi, ale design vypada jako typ (Osuška...), pouzij ho
        if not product_name and design_name and design_name.lower() in ("osuška", "osuska", "župan", "zupan", "podprsenka", "kalhotky"):
            product_name = design_name

        # Check if we have the model in products cache
        in_db = model_code in products_db
        
        rows.append({
            "row_num": r,
            "raw_code": code_str,
            "model_code": model_code,
            "color_code": color_code,
            "color_name": color_name,
            "arguments": args,
            "product_name": product_name,
            "design_name": design_name,
            "brand": brand,
            "material": material,
            "size": size,
            "in_db": in_db
        })
        
    logging.info(f"Získáno {len(rows)} datových řádků k zpracování.")
    return rows

def write_row_results_to_excel(file_path, row_num, results):
    """
    Writes the 12 generated copywriting results (CZ + SK) into Excel columns.
    If the columns do not exist in the sheet, they are added at the end of the header row.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    header_row_idx, mapping = find_header_row_and_mapping(ws)
    
    # Read current headers from the detected header row
    headers = [str(ws.cell(row=header_row_idx, column=c).value).strip() if ws.cell(row=header_row_idx, column=c).value is not None else "" for c in range(1, ws.max_column + 1)]
    
    col_write_map = {}
    
    # Define mapping of result keys to headers we look for or create
    key_to_headers = {
        "eshop_name": ["E-SHOP NÁZEV", "Název E-shop", "E-shop název"],
        "short_name": ["E-SHOP KRÁTKÝ NÁZEV", "E-shop krátký název", "Krátký název"],
        "eshop_desc1": ["TRIOLA ESHOP POPIS", "Dlouhý popis (HTML)", "Popis 1"],
        "eshop_desc2": ["TRIOLA ESHOP POPIS 2", "Popis 2"],
        "meta_title": ["ESHOP META TITLE", "Meta Title", "Meta title"],
        "meta_desc": ["ESHOP META DESCRIPTION", "Meta Description", "Meta description"],
        # slovenska mutace
        "eshop_name_sk": ["E-SHOP NÁZEV SK", "E-shop název SK"],
        "short_name_sk": ["E-SHOP KRÁTKÝ NÁZEV SK", "E-shop krátký název SK"],
        "eshop_desc1_sk": ["TRIOLA ESHOP POPIS SK", "Popis 1 SK"],
        "eshop_desc2_sk": ["TRIOLA ESHOP POPIS 2 SK", "Popis 2 SK"],
        "meta_title_sk": ["ESHOP META TITLE SK", "Meta Title SK"],
        "meta_desc_sk": ["ESHOP META DESCRIPTION SK", "Meta Description SK"],
    }
    
    def _is_sk_header(h):
        """Konci hlavicka jazykovou mutaci SK?"""
        hl = h.lower().strip()
        return hl.endswith(" sk") or hl.endswith("_sk") or hl.endswith("-sk")

    # Find existing columns (case-insensitive). CZ klic NESMI sebrat SK sloupec a naopak.
    used_cols = set()
    for key, possible_headers in key_to_headers.items():
        want_sk = key.endswith("_sk")
        found_idx = -1

        # 1. Exact match
        for idx, h in enumerate(headers):
            if (idx + 1) in used_cols or _is_sk_header(h) != want_sk:
                continue
            h_lower = h.lower().strip()
            if any(ph.lower().strip() == h_lower for ph in possible_headers):
                found_idx = idx + 1
                break

        # 2. Substring match (jen v ramci stejne jazykove mutace)
        if found_idx == -1:
            for idx, h in enumerate(headers):
                if (idx + 1) in used_cols or _is_sk_header(h) != want_sk:
                    continue
                h_lower = h.lower().strip()
                if not h_lower:
                    continue
                if any(ph.lower().strip() in h_lower or h_lower in ph.lower().strip() for ph in possible_headers):
                    found_idx = idx + 1
                    break

        if found_idx != -1:
            used_cols.add(found_idx)
        col_write_map[key] = found_idx

    # If any column is not found, append it to the header row
    for key, possible_headers in key_to_headers.items():
        if col_write_map[key] == -1:
            new_col_idx = len(headers) + 1
            default_header_name = possible_headers[0]
            ws.cell(row=header_row_idx, column=new_col_idx, value=default_header_name)
            headers.append(default_header_name)
            col_write_map[key] = new_col_idx
            
    # Write values
    for key, val in results.items():
        if key in col_write_map and col_write_map[key] != -1:
            col_idx = col_write_map[key]
            ws.cell(row=row_num, column=col_idx, value=val)
        
    wb.save(file_path)
    logging.info(f"Zapsáno {sum(1 for k in results if col_write_map.get(k, -1) != -1)} sloupců výsledků na řádek {row_num} v {file_path}")

def write_descriptions_to_excel(file_path, row_num, short_html, long_html):
    """
    Wrapper for backward compatibility.
    """
    results = {
        "short_name": short_html,
        "eshop_desc1": long_html
    }
    write_row_results_to_excel(file_path, row_num, results)

def normalize_header(h):
    if not h:
        return ""
    trans = str.maketrans("áéěíóúůýžščřďťň", "aeeouuyzscrdtnn")
    h = str(h).lower().strip().translate(trans)
    h = re.sub(r'[\s_\-]+', '', h)
    return h

def detect_csv_encoding_and_delimiter(file_path):
    encodings = ['utf-8-sig', 'windows-1250', 'utf-8', 'latin1']
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc) as f:
                first_line = f.readline()
                if not first_line:
                    continue
                semicolons = first_line.count(';')
                commas = first_line.count(',')
                tabs = first_line.count('\t')
                
                delimiter = ';'
                if commas > semicolons and commas > tabs:
                    delimiter = ','
                elif tabs > semicolons and tabs > commas:
                    delimiter = '\t'
                return enc, delimiter
        except UnicodeDecodeError:
            continue
    return 'utf-8', ';'

def parse_seo_batch(file_path):
    """
    Parses CSV or Excel file for SEO Snippet generation.
    Returns a list of dicts with normalized keys.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Soubor {file_path} neexistuje.")

    ext = os.path.splitext(file_path.lower())[1]
    rows = []
    headers = []

    if ext == '.csv':
        enc, delimiter = detect_csv_encoding_and_delimiter(file_path)
        with open(file_path, 'r', encoding=enc) as f:
            reader = csv.reader(f, delimiter=delimiter)
            try:
                headers = next(reader)
            except StopIteration:
                return []
            
            headers = [h.strip() for h in headers]
            
            for idx, row in enumerate(reader, start=2):
                if not row or all(cell.strip() == "" for cell in row):
                    continue
                rows.append((idx, row))
    elif ext in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        
        for idx in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=idx, column=c).value for c in range(1, ws.max_column + 1)]
            row_str_vals = [str(cell).strip() if cell is not None else "" for cell in row_vals]
            if all(cell == "" for cell in row_str_vals):
                continue
            rows.append((idx, row_str_vals))
    else:
        raise ValueError(f"Nepodporovaný formát souboru: {ext}")

    mapping = {}
    normalized_headers = [normalize_header(h) for h in headers]
    
    keyword_patterns = ['primarnikw', 'primarniklicoveslovo', 'klicoveslovo', 'klicovyvyraz', 'klicovydotaz', 'dotaz', 'kw']
    intent_patterns = ['intent', 'zamer', 'typzameru']
    type_patterns = ['typstranky', 'kategorie', 'pagetype', 'typ']
    title_patterns = ['soucasnytitle', 'aktualnititle', 'staritle', 'title']
    desc_patterns = ['soucasnadesc', 'soucasnydescription', 'aktualnidesc', 'description', 'desc', 'popisek']
    h1_patterns = ['soucasnyh1', 'aktualnih1', 'h1']
    usp_patterns = ['usp', 'vyhody', 'prednosti', 'argumenty']
    brand_patterns = ['brand', 'znacka', 'firma']
    serp_patterns = ['konkurenceserp', 'serpkonkurence', 'konkurence', 'titleskonkurence']
    url_patterns = ['url', 'adresa', 'link', 'stranka']

    def find_idx(patterns):
        for pattern in patterns:
            for i, nh in enumerate(normalized_headers):
                if nh == pattern:
                    return i
        for pattern in patterns:
            for i, nh in enumerate(normalized_headers):
                if pattern in nh or nh in pattern:
                    return i
        return -1

    mapping['url'] = find_idx(url_patterns)
    mapping['primarni_kw'] = find_idx(keyword_patterns)
    mapping['intent'] = find_idx(intent_patterns)
    mapping['typ_stranky'] = find_idx(type_patterns)
    mapping['soucasny_title'] = find_idx(title_patterns)
    mapping['soucasna_desc'] = find_idx(desc_patterns)
    mapping['soucasny_h1'] = find_idx(h1_patterns)
    mapping['usp'] = find_idx(usp_patterns)
    mapping['brand'] = find_idx(brand_patterns)
    mapping['konkurence_serp'] = find_idx(serp_patterns)

    standard_keys = ['url', 'primarni_kw', 'intent', 'typ_stranky', 'soucasny_title', 'soucasna_desc', 'soucasny_h1', 'usp', 'brand', 'konkurence_serp']
    for idx, key in enumerate(standard_keys):
        if mapping[key] == -1 and idx < len(headers):
            mapping[key] = idx

    parsed_data = []
    for idx, row in rows:
        def get_val(key, default=""):
            col_idx = mapping[key]
            if col_idx != -1 and col_idx < len(row):
                return str(row[col_idx]).strip()
            return default

        parsed_data.append({
            "row_num": idx,
            "url": get_val('url'),
            "primarni_kw": get_val('primarni_kw'),
            "intent": get_val('intent'),
            "typ_stranky": get_val('typ_stranky'),
            "soucasny_title": get_val('soucasny_title'),
            "soucasna_desc": get_val('soucasna_desc'),
            "soucasny_h1": get_val('soucasny_h1'),
            "usp": get_val('usp'),
            "brand": get_val('brand', 'Triola'),
            "konkurence_serp": get_val('konkurence_serp')
        })
    
    return parsed_data

def write_seo_to_file(file_path, row_num, seo_data):
    """
    Writes the generated SEO snippet data into new columns in the CSV/Excel file.
    Creates column headers if they don't exist yet.
    """
    ext = os.path.splitext(file_path.lower())[1]
    
    headers_to_add = [
        "Navržený Title",
        "Délka Title",
        "Navržený Description",
        "Délka Description",
        "Navržený H1",
        "Pattern Break",
        "Otevřená smyčka",
        "Rizika"
    ]
    
    vals_to_add = [
        seo_data.get("title", ""),
        seo_data.get("title_znaku", len(seo_data.get("title", ""))),
        seo_data.get("description", ""),
        seo_data.get("desc_znaku", len(seo_data.get("description", ""))),
        seo_data.get("h1", ""),
        seo_data.get("pattern_break", ""),
        seo_data.get("smycka", ""),
        seo_data.get("rizika", "")
    ]
    
    if ext == '.csv':
        enc, delimiter = detect_csv_encoding_and_delimiter(file_path)
        all_rows = []
        with open(file_path, 'r', encoding=enc) as f:
            reader = csv.reader(f, delimiter=delimiter)
            all_rows = list(reader)
            
        if not all_rows:
            return
            
        headers = all_rows[0]
        col_idxs = {}
        for h in headers_to_add:
            if h in headers:
                col_idxs[h] = headers.index(h)
            else:
                headers.append(h)
                col_idxs[h] = len(headers) - 1
                
        for r_idx in range(len(all_rows)):
            while len(all_rows[r_idx]) < len(headers):
                all_rows[r_idx].append("")
                
        target_idx = row_num - 1
        if 0 < target_idx < len(all_rows):
            for h, val in zip(headers_to_add, vals_to_add):
                col_i = col_idxs[h]
                all_rows[target_idx][col_i] = str(val)
                
        with open(file_path, 'w', encoding=enc, newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(all_rows)
            
    elif ext in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        
        col_idxs = {}
        for h in headers_to_add:
            if h in headers:
                col_idxs[h] = headers.index(h) + 1
            else:
                new_col = len(headers) + 1
                ws.cell(row=1, column=new_col, value=h)
                headers.append(h)
                col_idxs[h] = new_col
                
        for h, val in zip(headers_to_add, vals_to_add):
            col_i = col_idxs[h]
            ws.cell(row=row_num, column=col_i, value=val)
            
        wb.save(file_path)
        
    logging.info(f"Zapsána SEO data na řádek {row_num} v {file_path}")

