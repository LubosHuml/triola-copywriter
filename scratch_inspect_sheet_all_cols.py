import openpyxl

file_path = "triola_marketing_data.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Triola stálá kolekce"]

# Print all headers (Row 1)
row1 = [cell.value for cell in ws[1]]
print("Total columns in Triola stálá kolekce:", len(row1))
for col_idx, header in enumerate(row1, 1):
    print(f"Col {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {header}")
    
# Let's inspect some rows, focusing on column B (Nomenklatura) and columns containing arguments
print("\nSample Data rows:")
for r in range(2, 10):
    nomen = ws.cell(row=r, column=2).value # Nomenklatura
    name = ws.cell(row=r, column=6).value # E-SHOP NÁZEV
    desc = ws.cell(row=r, column=7).value # TRIOLA ESHOP POPIS
    desc2 = ws.cell(row=r, column=8).value # TRIOLA ESHOP POPIS 2
    
    # Check if there are other columns, e.g. columns 12, 13, 14, 15...
    extra_vals = []
    for c in range(12, min(len(row1) + 1, 25)):
        val = ws.cell(row=r, column=c).value
        if val:
            extra_vals.append(f"{row1[c-1]}: {str(val)[:30]}")
            
    print(f"\nRow {r}: Nomenklatura={nomen} | Eshop Název={name}")
    print(f"  Description 1: {str(desc)[:60] if desc else 'None'}")
    print(f"  Description 2: {str(desc2)[:60] if desc2 else 'None'}")
    if extra_vals:
        print(f"  Extra data: {', '.join(extra_vals)}")
