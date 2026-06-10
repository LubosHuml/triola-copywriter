import openpyxl
import re
import os
import json
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

EXCEL_FILE = "triola_marketing_data.xlsx"
MARKETING_CACHE_FILE = "marketing_cache.json"

def download_excel_sheet():
    """Downloads the Excel sheet from Google Drive direct download URL."""
    sheet_id = "1WO2WWAQrgb2PvpHkokl4r6TpWEI36mCC"
    download_url = f"https://drive.google.com/uc?export=download&id={sheet_id}"
    logging.info(f"Stahování marketingového Excelu z Google Drive ({download_url})...")
    try:
        import urllib.request
        req = urllib.request.Request(
            download_url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) TriolaCopywriter/1.0'}
        )
        with urllib.request.urlopen(req) as response:
            with open(EXCEL_FILE, "wb") as f:
                f.write(response.read())
        logging.info("Marketingový Excel úspěšně stažen a uložen.")
        return True
    except Exception as e:
        logging.error(f"Chyba při stahování marketingového Excelu: {e}")
        return False

def extract_model_code(value):
    """Safely extracts 5-digit model code from Excel cells (handling floats, strings, nomenclatures)."""
    if value is None:
        return None
    
    val_str = str(value).strip()
    
    # Handle excel float format (e.g. 28746.0)
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
        
    # Search for a 5-digit number with lookahead/lookbehind to prevent matching part of a longer digit sequence
    match = re.search(r'(?<!\d)\d{5}(?!\d)', val_str)
    if match:
        return match.group(0)
        
    # Fallback: remove non-digits and check if it's a 5-digit number
    digits = re.sub(r'\D', '', val_str)
    if len(digits) == 5:
        return digits
        
    return None

def build_marketing_db(force_update=False):
    """Parses Excel file and builds a marketing knowledge base mapped by model_code."""
    if force_update or not os.path.exists(EXCEL_FILE):
        download_excel_sheet()
        
    if not force_update and os.path.exists(MARKETING_CACHE_FILE):
        try:
            with open(MARKETING_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                logging.info(f"Načteno {len(data)} marketingových záznamů z cache.")
                return data
        except Exception as e:
            logging.error(f"Chyba při čtení marketing cache: {e}. Proběhne regenerace.")

    if not os.path.exists(EXCEL_FILE):
        logging.warning(f"Soubor {EXCEL_FILE} neexistuje. Aplikace poběží bez marketingových podkladů z Excelu.")
        return {}

    logging.info("Parsování marketingových podkladů z Excel tabulky...")
    marketing_db = {}

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # We focus on sheets containing "Triola"
        triola_sheets = [s for s in wb.sheetnames if "triola" in s.lower()]
        
        for sheet_name in triola_sheets:
            logging.info(f"Zpracování listu: {sheet_name}")
            ws = wb[sheet_name]
            
            # Identify columns
            # Some sheets have headers on row 1, some on row 2
            header_row = 1
            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[1]]
            
            # If row 1 is mostly empty, check row 2
            non_empty_headers = [h for h in headers if h]
            if len(non_empty_headers) < 3 and ws.max_row > 1:
                header_row = 2
                headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[2]]

            # Map columns to indices (0-indexed)
            col_map = {
                "code": -1,
                "cut": -1,
                "collection": -1,
                "arguments": -1,
                "target_group": -1,
                "description1": -1,
                "description2": -1,
                "meta_title": -1,
                "meta_desc": -1,
            }

            for idx, h in enumerate(headers):
                if not h:
                    continue
                # Code mapping
                if "fazón" in h or "nomenklatura" in h or "číslo" in h or h == "kod":
                    col_map["code"] = idx
                # Cut mapping
                elif "střih" in h or "strih" in h:
                    col_map["cut"] = idx
                # Collection mapping
                elif "kolekce" in h:
                    col_map["collection"] = idx
                # Arguments mapping
                elif "argument" in h or "prodejní" in h:
                    col_map["arguments"] = idx
                # Target group mapping
                elif "poprsí" in h or "věk" in h or "cílová" in h or "cilova" in h:
                    col_map["target_group"] = idx
                # Description mapping
                elif "eshop popis 2" in h or "popis 2" in h:
                    col_map["description2"] = idx
                elif "eshop popis" in h or "popis" in h:
                    col_map["description1"] = idx
                # Meta tags mapping
                elif "meta title" in h or "meta_title" in h or h == "meta title":
                    col_map["meta_title"] = idx
                elif "meta desc" in h or "meta_desc" in h or "meta description" in h:
                    col_map["meta_desc"] = idx

            # Log column mapping details
            logging.info(f" - Mapování sloupců: { {k: v for k, v in col_map.items() if v != -1} }")
            
            # If we don't have code column, skip sheet
            if col_map["code"] == -1:
                logging.warning(f" - List '{sheet_name}' přeskočen: Nenalezen sloupec pro kód produktu.")
                continue

            # Iterate over data rows
            start_row = header_row + 1
            for r in range(start_row, ws.max_row + 1):
                raw_code = ws.cell(row=r, column=col_map["code"] + 1).value
                model_code = extract_model_code(raw_code)
                
                if not model_code:
                    continue

                # Read fields
                cut = ws.cell(row=r, column=col_map["cut"] + 1).value if col_map["cut"] != -1 else None
                collection = ws.cell(row=r, column=col_map["collection"] + 1).value if col_map["collection"] != -1 else None
                arguments = ws.cell(row=r, column=col_map["arguments"] + 1).value if col_map["arguments"] != -1 else None
                target_group = ws.cell(row=r, column=col_map["target_group"] + 1).value if col_map["target_group"] != -1 else None
                desc1 = ws.cell(row=r, column=col_map["description1"] + 1).value if col_map["description1"] != -1 else None
                desc2 = ws.cell(row=r, column=col_map["description2"] + 1).value if col_map["description2"] != -1 else None
                meta_title = ws.cell(row=r, column=col_map["meta_title"] + 1).value if col_map["meta_title"] != -1 else None
                meta_desc = ws.cell(row=r, column=col_map["meta_desc"] + 1).value if col_map["meta_desc"] != -1 else None

                # Clean cell string values
                def clean_val(val):
                    if val is None: return ""
                    val_str = str(val).strip()
                    # Strip standard 'x' or '/' placeholders
                    if val_str.lower() in ('x', '/', '-'):
                        return ""
                    return val_str

                cut = clean_val(cut)
                collection = clean_val(collection)
                if not collection and "stálá kolekce" in sheet_name.lower():
                    collection = "Stálá kolekce"
                arguments = clean_val(arguments)
                target_group = clean_val(target_group)
                desc1 = clean_val(desc1)
                desc2 = clean_val(desc2)
                meta_title = clean_val(meta_title)
                meta_desc = clean_val(meta_desc)

                if model_code not in marketing_db:
                    marketing_db[model_code] = {
                        "model_code": model_code,
                        "cuts": [],
                        "collections": [],
                        "sales_arguments": [],
                        "target_groups": [],
                        "descriptions": [],
                        "meta_titles": [],
                        "meta_descriptions": [],
                        "sources": []
                    }

                entry = marketing_db[model_code]

                # Append non-duplicate values
                if cut and cut not in entry["cuts"]: entry["cuts"].append(cut)
                if collection and collection not in entry["collections"]: entry["collections"].append(collection)
                if arguments and arguments not in entry["sales_arguments"]: entry["sales_arguments"].append(arguments)
                if target_group and target_group not in entry["target_groups"]: entry["target_groups"].append(target_group)
                if desc1 and desc1 not in entry["descriptions"]: entry["descriptions"].append(desc1)
                if desc2 and desc2 not in entry["descriptions"]: entry["descriptions"].append(desc2)
                if meta_title and meta_title not in entry["meta_titles"]: entry["meta_titles"].append(meta_title)
                if meta_desc and meta_desc not in entry["meta_descriptions"]: entry["meta_descriptions"].append(meta_desc)
                if sheet_name not in entry["sources"]: entry["sources"].append(sheet_name)

        # Save compiled database to JSON cache
        with open(MARKETING_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(marketing_db, f, ensure_ascii=False, indent=2)
            
        logging.info(f"Zpracováno {len(marketing_db)} unikátních modelů z Excelu a uloženo do cache.")
        return marketing_db

    except Exception as e:
        logging.error(f"Chyba při parsování Excel souboru: {e}")
        return {}

def get_marketing_data_for_model(model_code, marketing_db):
    """Retrieves merged marketing parameters from Excel for a given model code."""
    if not model_code or not marketing_db:
        return {}
        
    code = str(model_code).strip()
    if code in marketing_db:
        entry = marketing_db[code]
        return {
            "collection": ", ".join(entry["collections"]),
            "sales_arguments": "\\n - " + "\\n - ".join(entry["sales_arguments"]) if entry["sales_arguments"] else "",
            "target_group": ", ".join(entry["target_groups"]),
            "meta_title": entry["meta_titles"][0] if entry["meta_titles"] else "",
            "meta_description": entry["meta_descriptions"][0] if entry["meta_descriptions"] else "",
            "extra_descriptions": "\\n".join([re.sub(r'<[^>]*>', '', d).strip() for d in entry["descriptions"]]) if entry["descriptions"] else ""
        }
    return {}

if __name__ == "__main__":
    print("Testování parseru marketingových tabulek...")
    db = build_marketing_db(force_update=True)
    
    # Test specific lookups
    test_codes = ["28746", "22884", "21864"]
    for code in test_codes:
        data = get_marketing_data_for_model(code, db)
        print(f"\nModel {code}:")
        for k, v in data.items():
            print(f"  {k}: {v[:100]}...")
