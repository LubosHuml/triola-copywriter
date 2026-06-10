import openpyxl

file_path = "triola_marketing_data.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in ["Triola JL 25", "Triola PZ 25"]:
    ws = wb[sheet_name]
    # Header is at row 1 or row 2?
    # Let's inspect rows 1 and 2
    row1 = [cell.value for cell in ws[1]]
    row2 = [cell.value for cell in ws[2]]
    
    print(f"\n======================================")
    print(f"SHEET: {sheet_name}")
    print(f"======================================")
    print("Row 1 headers:")
    for idx, r in enumerate(row1, 1):
        if r:
            print(f"  {idx} ({openpyxl.utils.get_column_letter(idx)}): {r}")
    print("\nRow 2 headers:")
    for idx, r in enumerate(row2, 1):
        if r:
            print(f"  {idx} ({openpyxl.utils.get_column_letter(idx)}): {r}")
            
    # Print some data rows
    print("\nSample Rows:")
    count = 0
    for r in range(3, 15):
        row_vals = [cell.value for cell in ws[r]]
        # Check if the row contains values
        if any(row_vals):
            # Try to map columns: fazona (model code) is usually Col 2, střih is usually Col 9, color is Col 10
            # Let's print the entire row (truncated cell values)
            cols_val = {openpyxl.utils.get_column_letter(idx): str(val)[:30] if val is not None else "" for idx, val in enumerate(row_vals, 1)}
            # Remove empty values to make it compact
            compact = {k: v for k, v in cols_val.items() if v}
            print(f"  Row {r}: {compact}")
            count += 1
            if count >= 5:
                break
